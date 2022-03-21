from pyexpat import model
from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from django.views.generic.edit import (CreateView, UpdateView, DeleteView)
from django.views.generic.base import TemplateView
from django.views import View
from openpyxl import load_workbook, Workbook
import secrets
from .models import *


class UrlListView(LoginRequiredMixin, ListView):
    template_name = 'attendance/url-list.html'
    model = AttendanceURLToken


class AddAttendanceView(LoginRequiredMixin, TemplateView):
    template_name = 'attendance/add-attendance.html'

    def post(self, request, *args, **kwargs):
        name = request.POST.get('name')
        remark = request.POST.get('remark')
        if name != None and name != '':
            filename = 'apps/attendance/static/attendance/'+kwargs['pk']+'.xlsx'
            wb = load_workbook(filename=filename)
            sheet = wb.active
            total_rows = sheet.max_row
            sheet.cell(row=total_rows, column=1).value = name
            sheet.cell(row=total_rows, column=2).value = remark
            wb.save(filename=filename)

        return redirect(reverse_lazy('attendance:url_list'))

class DeleteAttendanceUrlAndSheet(LoginRequiredMixin, DeleteView):
    template_name = 'attendance/url_confirm_delete.html'
    model = AttendanceURLToken
    success_url = reverse_lazy('attendance:url_list')